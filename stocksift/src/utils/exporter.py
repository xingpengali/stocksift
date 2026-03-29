# -*- coding: utf-8 -*-
"""
数据导出模块

提供 Excel、CSV 等格式的数据导出功能
"""
import csv
import io
from datetime import datetime, date
from decimal import Decimal
from typing import List, Dict, Any, Optional, Union
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from utils.logger import get_logger

logger = get_logger(__name__)


class ExportError(Exception):
    """导出错误异常"""
    pass


class DataExporter:
    """
    数据导出器
    
    支持多种格式的数据导出
    """
    
    def __init__(self):
        self._formats = {
            'csv': self._export_csv,
            'json': self._export_json,
        }
        
        if EXCEL_AVAILABLE:
            self._formats['excel'] = self._export_excel
            self._formats['xlsx'] = self._export_excel
    
    def export(self, data: List[Dict[str, Any]], 
               columns: Optional[List[tuple]] = None,
               format_type: str = 'csv',
               filename: Optional[str] = None) -> Union[str, bytes]:
        """
        导出数据
        
        Args:
            data: 数据列表，每项为字典
            columns: 列定义 [(字段名, 显示名), ...]
            format_type: 导出格式 (csv, excel, json)
            filename: 文件名（用于Excel）
            
        Returns:
            导出内容（字符串或字节）
            
        Raises:
            ExportError: 导出失败时抛出
        """
        if not data:
            raise ExportError("没有数据可导出")
        
        if format_type not in self._formats:
            raise ExportError(f"不支持的导出格式: {format_type}")
        
        # 自动推断列定义
        if columns is None:
            columns = [(k, k) for k in data[0].keys()]
        
        return self._formats[format_type](data, columns, filename)
    
    def _export_csv(self, data: List[Dict[str, Any]], 
                    columns: List[tuple], filename: Optional[str] = None) -> str:
        """
        导出为 CSV
        
        Args:
            data: 数据列表
            columns: 列定义
            filename: 文件名（可选）
            
        Returns:
            CSV 内容字符串
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        headers = [col[1] for col in columns]
        writer.writerow(headers)
        
        # 写入数据
        for row in data:
            row_data = []
            for field, _ in columns:
                value = row.get(field, '')
                # 处理特殊类型
                if isinstance(value, (datetime, date)):
                    value = value.strftime('%Y-%m-%d')
                elif isinstance(value, Decimal):
                    value = float(value)
                elif value is None:
                    value = ''
                row_data.append(value)
            writer.writerow(row_data)
        
        return output.getvalue()
    
    def _export_excel(self, data: List[Dict[str, Any]],
                      columns: List[tuple], filename: Optional[str] = None) -> bytes:
        """
        导出为 Excel
        
        Args:
            data: 数据列表
            columns: 列定义
            filename: 文件名（可选）
            
        Returns:
            Excel 文件字节内容
        """
        if not EXCEL_AVAILABLE:
            raise ExportError("openpyxl 模块未安装，无法导出 Excel")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 设置表头样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        for col_idx, (field, header) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, (field, _) in enumerate(columns, 1):
                value = row_data.get(field, '')
                
                # 处理特殊类型
                if isinstance(value, (datetime, date)):
                    value = value.strftime('%Y-%m-%d')
                elif isinstance(value, Decimal):
                    value = float(value)
                elif value is None:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 自动调整列宽
        for col_idx, (field, header) in enumerate(columns, 1):
            max_length = len(str(header))
            for row_data in data:
                value = row_data.get(field, '')
                max_length = max(max_length, len(str(value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _export_json(self, data: List[Dict[str, Any]],
                     columns: List[tuple], filename: Optional[str] = None) -> str:
        """
        导出为 JSON
        
        Args:
            data: 数据列表
            columns: 列定义
            filename: 文件名（可选）
            
        Returns:
            JSON 字符串
        """
        import json
        
        # 只保留指定列
        filtered_data = []
        for row in data:
            filtered_row = {}
            for field, header in columns:
                value = row.get(field)
                # 处理特殊类型
                if isinstance(value, (datetime, date)):
                    value = value.strftime('%Y-%m-%d')
                elif isinstance(value, Decimal):
                    value = float(value)
                filtered_row[header] = value
            filtered_data.append(filtered_row)
        
        return json.dumps(filtered_data, ensure_ascii=False, indent=2)
    
    def save_to_file(self, content: Union[str, bytes], filepath: str) -> None:
        """
        保存内容到文件
        
        Args:
            content: 内容（字符串或字节）
            filepath: 文件路径
        """
        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        
        mode = 'wb' if isinstance(content, bytes) else 'w'
        encoding = None if isinstance(content, bytes) else 'utf-8'
        
        with open(filepath, mode, encoding=encoding) as f:
            f.write(content)
        
        logger.info(f"数据已导出到: {filepath}")


class StockExporter:
    """
    股票数据专用导出器
    
    提供股票相关数据的便捷导出
    """
    
    # 股票列表列定义
    STOCK_LIST_COLUMNS = [
        ('code', '股票代码'),
        ('name', '股票名称'),
        ('exchange', '交易所'),
        ('market_type', '市场类型'),
        ('industry_name', '行业'),
        ('list_date', '上市日期'),
    ]
    
    # 行情数据列定义
    QUOTE_COLUMNS = [
        ('code', '股票代码'),
        ('name', '股票名称'),
        ('price', '最新价'),
        ('change', '涨跌额'),
        ('change_pct', '涨跌幅%'),
        ('volume', '成交量'),
        ('amount', '成交额'),
        ('pe_ttm', '市盈率'),
        ('pb', '市净率'),
    ]
    
    # K线数据列定义
    KLINE_COLUMNS = [
        ('code', '股票代码'),
        ('trade_date', '日期'),
        ('open', '开盘价'),
        ('high', '最高价'),
        ('low', '最低价'),
        ('close', '收盘价'),
        ('volume', '成交量'),
        ('amount', '成交额'),
        ('change_pct', '涨跌幅%'),
    ]
    
    # 财务数据列定义
    FINANCIAL_COLUMNS = [
        ('code', '股票代码'),
        ('report_date', '报告期'),
        ('report_type', '报告类型'),
        ('revenue', '营业收入'),
        ('net_profit', '净利润'),
        ('eps', '每股收益'),
        ('roe', '净资产收益率%'),
        ('gross_margin', '毛利率%'),
        ('debt_ratio', '资产负债率%'),
    ]
    
    def __init__(self):
        self._exporter = DataExporter()
    
    def export_stock_list(self, stocks: List[Dict], 
                          format_type: str = 'csv') -> Union[str, bytes]:
        """
        导出股票列表
        
        Args:
            stocks: 股票列表
            format_type: 导出格式
            
        Returns:
            导出内容
        """
        return self._exporter.export(
            stocks, 
            self.STOCK_LIST_COLUMNS, 
            format_type
        )
    
    def export_quotes(self, quotes: List[Dict],
                      format_type: str = 'csv') -> Union[str, bytes]:
        """
        导出行情数据
        
        Args:
            quotes: 行情列表
            format_type: 导出格式
            
        Returns:
            导出内容
        """
        return self._exporter.export(
            quotes,
            self.QUOTE_COLUMNS,
            format_type
        )
    
    def export_klines(self, klines: List[Dict],
                      format_type: str = 'csv') -> Union[str, bytes]:
        """
        导出 K 线数据
        
        Args:
            klines: K线列表
            format_type: 导出格式
            
        Returns:
            导出内容
        """
        return self._exporter.export(
            klines,
            self.KLINE_COLUMNS,
            format_type
        )
    
    def export_financials(self, financials: List[Dict],
                          format_type: str = 'csv') -> Union[str, bytes]:
        """
        导出财务数据
        
        Args:
            financials: 财务数据列表
            format_type: 导出格式
            
        Returns:
            导出内容
        """
        return self._exporter.export(
            financials,
            self.FINANCIAL_COLUMNS,
            format_type
        )
    
    def save_to_file(self, content: Union[str, bytes], filepath: str) -> None:
        """
        保存到文件
        
        Args:
            content: 内容
            filepath: 文件路径
        """
        self._exporter.save_to_file(content, filepath)


# 便捷函数
def export_to_csv(data: List[Dict], columns: Optional[List[tuple]] = None) -> str:
    """
    导出为 CSV 字符串
    
    Args:
        data: 数据列表
        columns: 列定义
        
    Returns:
        CSV 字符串
    """
    exporter = DataExporter()
    return exporter.export(data, columns, 'csv')


def export_to_excel(data: List[Dict], 
                    columns: Optional[List[tuple]] = None) -> bytes:
    """
    导出为 Excel 字节
    
    Args:
        data: 数据列表
        columns: 列定义
        
    Returns:
        Excel 字节内容
    """
    exporter = DataExporter()
    return exporter.export(data, columns, 'excel')
